import io
import re
import base64
import unicodedata
import pandas as pd
from flask import Flask, request, jsonify, render_template
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import json
import math

from logger import log_event

app = Flask(__name__, template_folder='templates')


class _SafeEncoder(json.JSONEncoder):
    """Converts NaN/Inf floats to None so the response is always valid JSON."""
    def default(self, obj):
        return super().default(obj)

    def iterencode(self, o, _one_shot=False):
        # Replace NaN and Infinity before encoding
        return super().iterencode(self._clean(o), _one_shot)

    def _clean(self, obj):
        if isinstance(obj, float):
            if math.isnan(obj) or math.isinf(obj):
                return None
        elif isinstance(obj, dict):
            return {k: self._clean(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [self._clean(v) for v in obj]
        return obj


app.json.encoder = _SafeEncoder


# ── File reading ───────────────────────────────────────────────────────────────

def read_file_from_bytes(file_bytes, filename, sheet_name=None):
    ext = filename.rsplit('.', 1)[1].lower()
    bio = io.BytesIO(file_bytes)
    if ext == 'csv':
        return pd.read_csv(bio, dtype=str)
    kwargs = {'dtype': str}
    if sheet_name:
        kwargs['sheet_name'] = sheet_name
    return pd.read_excel(bio, **kwargs)


def read_all_sheets_info(file_bytes, filename):
    ext = filename.rsplit('.', 1)[1].lower()
    if ext == 'csv':
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
        return {'(hoja única)': {'columns': list(df.columns), 'rows': len(df)}}
    try:
        xl  = pd.ExcelFile(io.BytesIO(file_bytes))
        out = {}
        for sn in xl.sheet_names:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sn, dtype=str)
            out[sn] = {'columns': list(df.columns), 'rows': len(df)}
        return out
    except Exception:
        return {}


def extract_header_styles(file_bytes, filename, sheet_name=None):
    ext = filename.rsplit('.', 1)[1].lower()
    if ext == 'csv':
        return {}
    styles = {}
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        for cell in ws[1]:
            col_name = str(cell.value) if cell.value is not None else ''
            if not col_name:
                continue
            style = {}
            try:
                if cell.fill and cell.fill.fill_type not in (None, 'none'):
                    fg = cell.fill.fgColor
                    if fg.type == 'rgb' and fg.rgb not in ('00000000', 'FFFFFFFF', '00FFFFFF'):
                        style['bg_color'] = fg.rgb
            except Exception:
                pass
            try:
                if cell.font:
                    style['bold'] = bool(cell.font.bold)
                    if cell.font.color and cell.font.color.type == 'rgb':
                        fc = cell.font.color.rgb
                        if fc not in ('FF000000', '00000000'):
                            style['font_color'] = fc
            except Exception:
                pass
            try:
                col_letter = cell.column_letter
                dim = ws.column_dimensions.get(col_letter)
                if dim and dim.width:
                    style['width'] = dim.width
            except Exception:
                pass
            styles[col_name] = style
        wb.close()
    except Exception:
        pass
    return styles


# ── Spell check helpers ────────────────────────────────────────────────────────

_PROPER_NAMES = [
    # Apellidos comunes con tilde o ñ
    'García', 'López', 'Martínez', 'González', 'Rodríguez', 'Hernández',
    'Pérez', 'Sánchez', 'Ramírez', 'Gómez', 'Díaz', 'Álvarez',
    'Gutiérrez', 'Jiménez', 'Muñoz', 'Domínguez', 'Núñez', 'Valdés',
    'Ríos', 'Suárez', 'Fernández', 'Cortés', 'Méndez', 'Vélez',
    'Benítez', 'Ibáñez', 'Córdoba', 'Zárate', 'Ordóñez', 'Céspedes',
    'Dávila', 'León', 'Guzmán', 'Vásquez', 'Peñaloza', 'Cárdenas',
    'Báez', 'Aráoz', 'Avilés', 'Aristizábal', 'Estévez', 'Páez',
    'Téllez', 'Chávez', 'Añez', 'Yáñez', 'Jáuregui', 'Gálvez',
    # Nombres propios con tilde o ñ
    'María', 'José', 'Ángel', 'Ramón', 'Andrés', 'Tomás', 'Víctor',
    'Álvaro', 'Héctor', 'Mónica', 'Néstor', 'Adrián', 'Sebastián',
    'Rubén', 'René', 'César', 'Nicolás', 'Damián', 'Fabián', 'Julián',
    'Germán', 'Simón', 'Agustín', 'Martín', 'Benjamín', 'Raúl', 'Saúl',
    'Óscar', 'Sofía', 'Verónica', 'Inés', 'Aníbal', 'Hernán',
    'Cristián', 'Valentín', 'Belén', 'Vivián', 'Érica', 'Úrsula',
    'Débora', 'Yésica', 'Mónica', 'Ángela', 'Azucena', 'Angélica',
    'Cintia', 'Mirián', 'Noé', 'Matías', 'Elías', 'Isaías', 'Tomás',
]

def _strip_accents(s):
    """Quita tildes y diéresis (ñ → n, á → a, etc.)."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

# Mapa: forma normalizada (sin tildes, minúsculas) → forma correcta con tildes
_PROPER_NOUN_FIXES = {_strip_accents(n).lower(): n for n in _PROPER_NAMES}

_spell_es = None

def _get_spell():
    global _spell_es
    if _spell_es is None:
        from spellchecker import SpellChecker
        _spell_es = SpellChecker(language='es')
    return _spell_es


def find_spelling_issues(df, columns, max_issues_per_col=50):
    """
    Detecta errores ortográficos en columnas de texto.
    - Palabras Title Case → revisión en _PROPER_NOUN_FIXES (tildes en nombres propios).
    - Palabras lowercase  → revisión con pyspellchecker (palabras comunes en español).
    Retorna lista de {column, issues: [{original_cell, word, suggested_word, count}]}
    """
    spell = _get_spell()
    result = []

    for col in columns:
        if col not in df.columns:
            continue
        series = df[col].dropna().astype(str)
        if series.empty:
            continue
        # Saltar columnas puramente numéricas
        if pd.to_numeric(series, errors='coerce').notna().mean() > 0.8:
            continue

        counts     = series.value_counts()
        col_issues = []
        seen_cells = set()

        for cell_val, cnt in counts.items():
            if cell_val in seen_cells or len(col_issues) >= max_issues_per_col:
                break

            # Tokenize: solo palabras (letras + apóstrofe), largo mínimo 3
            words = re.findall(r"[A-Za-záéíóúüÁÉÍÓÚÜñÑ']{3,}", cell_val)
            for word in words:
                letters = [c for c in word if c.isalpha()]
                if not letters:
                    continue

                # Rama 1: Title Case → revisión de nombre propio (solo tildes)
                if word[0].isupper() and not word.isupper():
                    key = _strip_accents(word).lower()
                    correct = _PROPER_NOUN_FIXES.get(key)
                    if correct and correct != word:
                        col_issues.append({
                            'original_cell': cell_val,
                            'word':          word,
                            'suggested_word': correct,
                            'count':         int(cnt),
                            'kind':          'nombre_propio',
                        })
                        seen_cells.add(cell_val)
                        break  # una corrección por celda

                # Rama 2: lowercase → revisión con pyspellchecker
                elif word.islower():
                    misspelled = spell.unknown([word])
                    if misspelled:
                        correction = spell.correction(word)
                        if correction and correction != word:
                            col_issues.append({
                                'original_cell': cell_val,
                                'word':          word,
                                'suggested_word': correction,
                                'count':         int(cnt),
                                'kind':          'comun',
                            })
                            seen_cells.add(cell_val)
                            break  # una corrección por celda

        if col_issues:
            result.append({'column': col, 'issues': col_issues})

    return result


# ── Analysis helpers ───────────────────────────────────────────────────────────

def normalize_string(s):
    """Lowercase, remove accents, remove punctuation, collapse spaces."""
    s = str(s).lower().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def find_similar_values(df, columns):
    """Return per-column groups of values that normalize to the same string."""
    suggestions = []
    for col in columns:
        if col not in df.columns:
            continue
        series = df[col].dropna().astype(str)
        if series.empty:
            continue
        # Skip purely numeric columns
        if pd.to_numeric(series, errors='coerce').notna().all():
            continue
        unique_vals = series.unique()
        if len(unique_vals) < 2 or len(unique_vals) > 300:
            continue

        # Group by normalized form
        norm_map = {}
        for v in unique_vals:
            key = normalize_string(v)
            norm_map.setdefault(key, []).append(v)

        col_groups = []
        for variants in norm_map.values():
            unique_variants = list(dict.fromkeys(variants))  # dedupe, preserve order
            if len(unique_variants) < 2:
                continue
            # Canonical = most frequent value
            counts = series.value_counts()
            canonical = max(unique_variants, key=lambda v: counts.get(v, 0))
            col_groups.append({'canonical': canonical, 'variants': unique_variants})

        if col_groups:
            suggestions.append({'column': col, 'groups': col_groups})
    return suggestions


def detect_casing_issue(value):
    """
    Detecta el tipo de problema de capitalización en un string.
    Retorna (issue_type, suggested) o (None, None) si no hay problema.
    """
    s = str(value).strip()
    # Saltar si es muy corto, numérico, email, URL o código
    if len(s) <= 1:
        return None, None
    if re.match(r'^[\d\s\-_/\\.,;:@#$%&*()\[\]{}]+$', s):
        return None, None
    if re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', s):  # email
        return None, None
    if re.match(r'^https?://', s):  # URL
        return None, None

    words = s.split()
    alpha_words = [w for w in words if any(c.isalpha() for c in w)]
    if not alpha_words:
        return None, None

    # Mezcla rara: tiene mayúsculas en posición inusual (ej: sANTIAGO, hOLA)
    def is_weird_mixed(w):
        letters = [c for c in w if c.isalpha()]
        if len(letters) < 3:
            return False
        # Si la primera letra es minúscula y hay mayúsculas después → raro
        if letters[0].islower() and any(c.isupper() for c in letters[1:]):
            return True
        # Si hay mayúsculas en el medio que no son la primera → raro
        if letters[0].isupper() and letters[1].isupper() and letters[-1].islower():
            # podría ser acrónimo, no marcar
            pass
        return False

    has_weird = any(is_weird_mixed(w) for w in alpha_words)
    all_upper = all(w.isupper() for w in alpha_words)
    all_lower = all(w.islower() for w in alpha_words)

    def _title_word(w):
        result, cap = [], True
        for ch in w:
            if ch.isalpha():
                result.append(ch.upper() if cap else ch.lower())
                cap = False
            else:
                result.append(ch)
                cap = True
        return ''.join(result)
    suggested = ' '.join(_title_word(w) for w in words)

    if has_weird:
        return 'mezcla_rara', suggested
    if all_upper and len(s) > 2:
        return 'todo_mayusculas', suggested
    if all_lower and len(alpha_words) >= 1:
        # Solo reportar minúsculas si parece nombre propio (más de una palabra
        # o empieza con minúscula siendo la primera palabra)
        if len(alpha_words) > 1 or (len(alpha_words) == 1 and len(alpha_words[0]) > 2):
            return 'todo_minusculas', suggested
    return None, None


def find_casing_issues(df, columns, max_issues_per_col=50):
    """
    Recorre columnas de texto y detecta problemas de capitalización.
    Retorna lista de {column, issues: [{original, suggested, issue_type, count}]}
    """
    result = []
    issue_labels = {
        'todo_mayusculas': 'TODO EN MAYÚSCULAS',
        'todo_minusculas': 'todo en minúsculas',
        'mezcla_rara':     'Mezcla rara de mayúsculas',
    }

    for col in columns:
        if col not in df.columns:
            continue
        series = df[col].dropna().astype(str)
        if series.empty:
            continue
        # Saltar columnas puramente numéricas
        if pd.to_numeric(series, errors='coerce').notna().mean() > 0.8:
            continue

        counts    = series.value_counts()
        col_issues = []
        seen       = set()

        for val, cnt in counts.items():
            if val in seen or len(col_issues) >= max_issues_per_col:
                break
            issue_type, suggested = detect_casing_issue(val)
            if issue_type and suggested and suggested != val:
                seen.add(val)
                col_issues.append({
                    'original':   val,
                    'suggested':  suggested,
                    'issue_type': issue_type,
                    'label':      issue_labels.get(issue_type, issue_type),
                    'count':      int(cnt),
                })

        if col_issues:
            result.append({'column': col, 'issues': col_issues})

    return result


def find_duplicate_groups(df, dedup_column, max_groups=50):
    """Return list of duplicate groups with their row ids and source files."""
    data_cols = [c for c in df.columns if c not in ('__source__', '__row_id__')]
    groups = []

    if dedup_column and dedup_column in df.columns:
        dup_mask = df.duplicated(subset=[dedup_column], keep=False)
        if not dup_mask.any():
            return []
        for key_val in df.loc[dup_mask, dedup_column].unique()[:max_groups]:
            mask    = df[dedup_column] == key_val
            grp_df  = df[mask]
            groups.append({
                'key_col': dedup_column,
                'key_val': str(key_val),
                'indices': grp_df['__row_id__'].tolist(),
                'records': grp_df[data_cols].fillna('').to_dict('records'),
                'sources': grp_df['__source__'].tolist() if '__source__' in df.columns else [],
            })
    else:
        # Full-row duplicates
        df_hash = df[data_cols].copy()
        df_hash['__hash__'] = df_hash.apply(
            lambda r: hash(tuple(str(x) for x in r.values)), axis=1)
        dup_mask = df_hash.duplicated(subset=['__hash__'], keep=False)
        if not dup_mask.any():
            return []
        seen = set()
        for _, row in df_hash[dup_mask].iterrows():
            h = row['__hash__']
            if h in seen or len(groups) >= max_groups:
                continue
            seen.add(h)
            same = df[df_hash['__hash__'] == h]
            groups.append({
                'key_col': None,
                'key_val': f'dup_{len(groups)}',
                'indices': same['__row_id__'].tolist(),
                'records': same[data_cols].fillna('').to_dict('records'),
                'sources': same['__source__'].tolist() if '__source__' in df.columns else [],
            })
    return groups


def _build_merged(file_data_list, column_order, sheet_selection, priority_file):
    """Common merge logic used by /preview and /confirm."""
    if priority_file:
        pri    = [fd for fd in file_data_list if fd['name'] == priority_file]
        others = [fd for fd in file_data_list if fd['name'] != priority_file]
        file_data_list = pri + others

    dfs = []
    for fd in file_data_list:
        file_bytes  = base64.b64decode(fd['data'])
        sheet_name  = sheet_selection.get(fd['name']) or None
        df          = read_file_from_bytes(file_bytes, fd['name'], sheet_name=sheet_name)
        cols_to_use = [c for c in column_order if c in df.columns]
        sub         = df[cols_to_use].copy()
        sub['__source__'] = fd['name']
        dfs.append(sub)

    merged = pd.concat(dfs, ignore_index=True)
    merged['__row_id__'] = range(len(merged))
    return merged


def _write_excel(merged_df, final_cols, column_styles):
    wb = Workbook()
    ws = wb.active

    for col_idx, col_name in enumerate(final_cols, 1):
        cell  = ws.cell(row=1, column=col_idx, value=col_name)
        style = column_styles.get(col_name, {})
        if 'bg_color' in style:
            cell.fill = PatternFill('solid', fgColor=style['bg_color'])
        font_kwargs = {'bold': style.get('bold', True)}
        if 'font_color' in style:
            font_kwargs['color'] = style['font_color']
        cell.font      = Font(**font_kwargs)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for row_idx, row in enumerate(merged_df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx,
                    value='' if (value is None or (isinstance(value, float) and pd.isna(value))) else value)

    for col_idx, col_name in enumerate(final_cols, 1):
        col_letter = get_column_letter(col_idx)
        style      = column_styles.get(col_name, {})
        ws.column_dimensions[col_letter].width = style.get('width', max(len(col_name) + 4, 14))

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return base64.b64encode(out.read()).decode('utf-8')


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files' not in request.files:
        return jsonify({'error': 'No se enviaron archivos'}), 400

    files       = request.files.getlist('files')
    all_styles  = {}
    file_data_list = []
    file_info   = []
    files_log   = []

    for f in files:
        if not f or not f.filename:
            continue
        ext = f.filename.rsplit('.', 1)[1].lower()
        if ext not in ('xlsx', 'xls', 'csv'):
            continue

        file_bytes = f.read()
        try:
            sheets_info   = read_all_sheets_info(file_bytes, f.filename)
            default_sheet = next(iter(sheets_info))

            file_data_list.append({
                'name': f.filename,
                'data': base64.b64encode(file_bytes).decode('utf-8'),
            })
            file_info.append({
                'name':          f.filename,
                'sheets':        sheets_info,
                'default_sheet': default_sheet,
            })
            files_log.append({'name': f.filename, 'size_kb': round(len(file_bytes) / 1024, 2)})

            styles = extract_header_styles(file_bytes, f.filename, sheet_name=default_sheet)
            for col, style in styles.items():
                if col not in all_styles:
                    all_styles[col] = style
        except Exception as e:
            return jsonify({'error': f'Error leyendo {f.filename}: {str(e)}'}), 400

    if not file_data_list:
        return jsonify({'error': 'Ningún archivo válido fue cargado'}), 400

    log_event(action='upload', ip=request.remote_addr, files=files_log, mode='unificar')

    return jsonify({
        'column_styles': all_styles,
        'files':         file_info,
        'file_data':     file_data_list,
    })


@app.route('/preview', methods=['POST'])
def preview():
    data            = request.get_json()
    column_order    = data.get('columns', [])
    dedup_column    = data.get('dedup_column') or None
    priority_file   = data.get('priority_file') or None
    sheet_selection = data.get('sheet_selection', {})
    file_data_list  = data.get('file_data', [])

    if not column_order or not file_data_list:
        return jsonify({'error': 'Faltan parámetros'}), 400

    try:
        merged = _build_merged(file_data_list, column_order, sheet_selection, priority_file)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

    final_cols = [c for c in column_order if c in merged.columns]

    similarity_suggestions = find_similar_values(merged[final_cols], final_cols)
    duplicate_groups       = find_duplicate_groups(merged, dedup_column)
    casing_issues          = find_casing_issues(merged, final_cols)
    spelling_issues        = find_spelling_issues(merged[final_cols], final_cols)
    truncated              = len(duplicate_groups) == 50

    return jsonify({
        'columns':               final_cols,
        'total_rows':            len(merged),
        'duplicate_groups':      duplicate_groups,
        'similarity_suggestions': similarity_suggestions,
        'casing_issues':         casing_issues,
        'spelling_issues':       spelling_issues,
        'truncated_duplicates':  truncated,
    })


@app.route('/confirm', methods=['POST'])
def confirm():
    data            = request.get_json()
    column_order    = data.get('columns', [])
    column_styles   = data.get('column_styles', {})
    dedup_column    = data.get('dedup_column') or None
    priority_file   = data.get('priority_file') or None
    sheet_selection = data.get('sheet_selection', {})
    file_data_list  = data.get('file_data', [])
    homologation    = data.get('homologation', {})    # {col: {old: new}}
    rows_to_delete  = set(data.get('rows_to_delete', []))
    shown_row_ids   = set(data.get('shown_row_ids', []))
    casing_fixes    = data.get('casing_fixes', {})    # {col: {original: suggested}}
    spelling_fixes  = data.get('spelling_fixes', {})  # {col: {original_cell: corrected_cell}}
    user_reviewed   = data.get('user_reviewed', False)

    if not column_order or not file_data_list:
        return jsonify({'error': 'Faltan parámetros'}), 400

    try:
        merged = _build_merged(file_data_list, column_order, sheet_selection, priority_file)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

    final_cols = [c for c in column_order if c in merged.columns]

    # Apply homologation
    for col, mapping in homologation.items():
        if col in merged.columns:
            merged[col] = merged[col].replace(mapping)

    # Apply casing fixes
    for col, mapping in casing_fixes.items():
        if col in merged.columns:
            merged[col] = merged[col].replace(mapping)

    # Apply spelling fixes
    for col, mapping in spelling_fixes.items():
        if col in merged.columns:
            merged[col] = merged[col].replace(mapping)

    before = len(merged)

    if user_reviewed:
        if rows_to_delete:
            merged = merged[~merged['__row_id__'].isin(rows_to_delete)]
        # Dedup rows not shown in the review UI (groups beyond the 50-group limit)
        if dedup_column and dedup_column in merged.columns:
            in_shown     = merged['__row_id__'].isin(shown_row_ids)
            shown_part   = merged[in_shown]
            unshown_part = merged[~in_shown].drop_duplicates(subset=[dedup_column], keep='first')
            merged = pd.concat([shown_part, unshown_part]).sort_values('__row_id__')
    else:
        # Fast path (no duplicates found) → apply default dedup
        if dedup_column and dedup_column in merged.columns:
            merged = merged.drop_duplicates(subset=[dedup_column], keep='first')
        else:
            merged = merged.drop_duplicates(
                subset=[c for c in final_cols if c in merged.columns], keep='first')

    duplicates_removed = before - len(merged)
    merged = merged[final_cols]

    output_b64 = _write_excel(merged, final_cols, column_styles)

    log_event(
        action='confirm',
        ip=request.remote_addr,
        files=[{'name': fd['name'], 'size_kb': 0} for fd in file_data_list],
        mode='unificar',
        rows=len(merged),
    )

    return jsonify({
        'output_data':        output_b64,
        'total_rows':         len(merged),
        'duplicates_removed': duplicates_removed,
        'columns':            final_cols,
    })


@app.route('/merge', methods=['POST'])
def merge_files():
    """Fast path — kept for backward compatibility."""
    data            = request.get_json()
    column_order    = data.get('columns', [])
    column_styles   = data.get('column_styles', {})
    dedup_column    = data.get('dedup_column') or None
    priority_file   = data.get('priority_file') or None
    sheet_selection = data.get('sheet_selection', {})
    file_data_list  = data.get('file_data', [])

    if not column_order or not file_data_list:
        return jsonify({'error': 'Faltan parámetros'}), 400

    try:
        merged = _build_merged(file_data_list, column_order, sheet_selection, priority_file)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

    final_cols = [c for c in column_order if c in merged.columns]
    merged     = merged[final_cols]

    before = len(merged)
    if dedup_column and dedup_column in merged.columns:
        merged = merged.drop_duplicates(subset=[dedup_column], keep='first')
    else:
        merged = merged.drop_duplicates(keep='first')

    output_b64 = _write_excel(merged, final_cols, column_styles)

    return jsonify({
        'output_data':        output_b64,
        'total_rows':         len(merged),
        'duplicates_removed': before - len(merged),
        'columns':            final_cols,
    })


# ── Formatting helpers ─────────────────────────────────────────────────────────

_PARTICLES = {'de', 'del', 'la', 'las', 'los', 'y', 'en', 'a', 'el', 'un', 'una', 'por', 'con'}


def apply_casing(value, rule):
    """Aplica transformación de texto según la regla."""
    if not isinstance(value, str):
        value = str(value)
    if rule == 'upper':
        return value.upper()
    if rule == 'lower':
        return value.lower()
    if rule == 'title':
        words = value.split()
        result = []
        for i, w in enumerate(words):
            if i > 0 and w.lower() in _PARTICLES:
                result.append(w.lower())
            else:
                result.append(w[0].upper() + w[1:].lower() if w else w)
        return ' '.join(result)
    return value


def clean_rut(val):
    """Limpia RUT: elimina puntos, guión y dígito verificador."""
    val = str(val).strip().replace('.', '').replace(' ', '')
    if '-' in val:
        val = val.rsplit('-', 1)[0]
    elif val:
        val = val[:-1]  # sin guion: el último caracter es el DV
    return re.sub(r'[^\d]', '', val)


_RUT_RE             = re.compile(r'^\d{1,2}\.?\d{3}\.?\d{3}-?[\dkK]$')
_EMAIL_RE           = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w+$')
_REGEX_EMAIL_STRICT = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')


def detect_column_type(series, col_name):
    """Detecta el tipo de columna: 'rut', 'email', 'text' u 'other'."""
    name_lower = col_name.lower()
    if 'rut' in name_lower or 'run' in name_lower:
        return 'rut'
    non_null = series.dropna().astype(str)
    if len(non_null) > 0:
        if non_null.apply(lambda x: bool(_RUT_RE.match(x.strip()))).mean() > 0.7:
            return 'rut'
        if non_null.apply(lambda x: bool(_EMAIL_RE.match(x.strip()))).mean() > 0.5:
            return 'email'
    if pd.to_numeric(series, errors='coerce').notna().mean() > 0.8:
        return 'other'
    if series.dtype == object:
        return 'text'
    return 'other'


def validate_emails(series):
    """
    Valida correos de una columna contra el regex estricto.
    Retorna lista de {row: N, value: "..."} donde N es número de fila Excel
    (encabezado = fila 1, primera dato = fila 2).
    """
    invalid = []
    for idx, val in series.items():
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        s = str(val).strip()
        if not s:
            continue
        if not _REGEX_EMAIL_STRICT.match(s):
            invalid.append({'row': int(idx) + 2, 'value': s})
    return invalid


def suggest_rule(detected_type):
    """Sugiere la regla de formato por defecto según el tipo de columna."""
    return {'rut': 'clean_rut', 'email': 'none', 'text': 'upper', 'other': 'none'}.get(detected_type, 'none')


def remove_accents(value):
    """Elimina tildes y diacríticos (á→a, é→e, ñ→n, ü→u, etc.)."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', str(value))
        if unicodedata.category(c) != 'Mn'
    )


# ── Formatting routes ──────────────────────────────────────────────────────────

@app.route('/format/preview', methods=['POST'])
def format_preview():
    """Analiza un archivo y retorna tipos de columna e hipervínculos detectados."""
    if request.content_type and 'multipart' in request.content_type:
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'error': 'No se envió archivo'}), 400
        ext = f.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls', 'csv'):
            return jsonify({'error': 'Formato no soportado'}), 400
        file_bytes    = f.read()
        filename      = f.filename
        sheet_name    = request.form.get('sheet_name') or None
        file_data_b64 = base64.b64encode(file_bytes).decode('utf-8')
    else:
        body          = request.get_json()
        file_data_b64 = body.get('file_data')
        filename      = body.get('filename', 'archivo.xlsx')
        sheet_name    = body.get('sheet_name') or None
        file_bytes    = base64.b64decode(file_data_b64)

    sheets_info = read_all_sheets_info(file_bytes, filename)
    if not sheet_name:
        sheet_name = next(iter(sheets_info))

    log_event(
        action='upload',
        ip=request.remote_addr,
        files=[{'name': filename, 'size_kb': round(len(file_bytes) / 1024, 2)}],
        mode='formatear',
    )

    # Detectar hipervínculos con openpyxl
    hyperlinks_by_col = {}
    total_hyperlinks  = 0
    ext = filename.rsplit('.', 1)[-1].lower()
    if ext != 'csv':
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            headers = [cell.value for cell in ws[1]]
            # Conteo total en una sola pasada (incluye columnas sin encabezado)
            total_hyperlinks = sum(
                1 for row in ws.iter_rows(min_row=2)
                for cell in row if cell.hyperlink is not None
            )
            # Conteo por columna para los badges de la tabla
            for col_idx, header in enumerate(headers, 1):
                if header is None:
                    continue
                count = sum(
                    1 for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx)
                    for cell in row if cell.hyperlink is not None
                )
                hyperlinks_by_col[str(header)] = count
            wb.close()
        except Exception:
            pass

    # Detectar tipos de columna con pandas
    try:
        df = read_file_from_bytes(file_bytes, filename, sheet_name=sheet_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

    columns = []
    for col in df.columns:
        col_str = str(col)
        dtype   = detect_column_type(df[col], col_str)
        col_info = {
            'name':             col_str,
            'detected_type':    dtype,
            'suggested_rule':   suggest_rule(dtype),
            'sample_values':    df[col].dropna().astype(str).head(3).tolist(),
            'hyperlinks_found': hyperlinks_by_col.get(col_str, 0),
        }
        # Validar correos inválidos en columnas de tipo email
        if dtype == 'email':
            invalid_rows = validate_emails(df[col])
            col_info['invalid_emails'] = len(invalid_rows)
            col_info['invalid_rows']   = invalid_rows
        columns.append(col_info)

    return jsonify({
        'sheets':             sheets_info,
        'sheet_name':         sheet_name,
        'columns':            columns,
        'file_data':          file_data_b64,
        'filename':           filename,
        'total_hyperlinks':   total_hyperlinks,
    })


@app.route('/format/confirm', methods=['POST'])
def format_confirm():
    """Aplica reglas de formato y retorna el archivo procesado."""
    body       = request.get_json()
    file_data  = body.get('file_data')
    filename   = body.get('filename', 'archivo.xlsx')
    sheet_name = body.get('sheet_name') or None
    rules      = body.get('rules', {})

    if not file_data:
        return jsonify({'error': 'Faltan datos del archivo'}), 400

    file_bytes        = base64.b64decode(file_data)
    ext               = filename.rsplit('.', 1)[-1].lower()
    email_corrections = body.get('email_corrections', {})  # {col: [{row, action, new_value?}]}

    # Construir estructuras de correcciones de correo (índice 0-based = row - 2)
    email_rows_to_delete = set()   # índices 0-based de filas a eliminar
    email_fixes_by_col   = {}      # {col_name: {idx_0based: new_value}}
    emails_fixed         = 0

    for col_name, corrections in email_corrections.items():
        for c in corrections:
            row_excel = c.get('row')
            action    = c.get('action', 'keep')
            if row_excel is None:
                continue
            idx = int(row_excel) - 2
            if action == 'delete':
                email_rows_to_delete.add(idx)
            elif action == 'fix':
                new_val = c.get('new_value', '')
                if new_val and _REGEX_EMAIL_STRICT.match(str(new_val).strip()):
                    email_fixes_by_col.setdefault(col_name, {})[idx] = str(new_val).strip()

    if ext == 'csv':
        df = read_file_from_bytes(file_bytes, filename)

        # Aplicar correcciones de correo antes del casing
        for col_name, fixes in email_fixes_by_col.items():
            if col_name not in df.columns:
                continue
            for idx, new_val in fixes.items():
                if idx in df.index:
                    df.at[idx, col_name] = new_val
                    emails_fixed += 1

        # Eliminar filas marcadas
        rows_removed = len([i for i in email_rows_to_delete if i in df.index])
        if email_rows_to_delete:
            df = df.drop(index=[i for i in email_rows_to_delete if i in df.index])

        for col_name, rule in rules.items():
            if rule == 'none' or col_name not in df.columns:
                continue
            if rule == 'clean_rut':
                df[col_name] = df[col_name].map(
                    lambda v: remove_accents(clean_rut(v)) if pd.notna(v) else v)
            else:
                df[col_name] = df[col_name].map(
                    lambda v, r=rule: remove_accents(apply_casing(str(v), r)) if pd.notna(v) else v)
        preview_cols = list(df.columns)
        preview_rows = [[str(v) for v in row] for row in df.head(5).fillna('').values.tolist()]
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(preview_cols)
        for row in df.itertuples(index=False):
            ws2.append(list(row))
        out = io.BytesIO()
        wb2.save(out)
        out.seek(0)
        log_event(
            action='confirm',
            ip=request.remote_addr,
            files=[{'name': filename, 'size_kb': 0}],
            mode='formatear',
            rows=len(df),
        )
        return jsonify({
            'output_data':        base64.b64encode(out.read()).decode('utf-8'),
            'hyperlinks_removed': 0,
            'emails_fixed':       emails_fixed,
            'rows_removed':       rows_removed,
            'preview_columns':    preview_cols,
            'preview_rows':       preview_rows,
        })

    # xlsx / xls: leer fuente con openpyxl, escribir resultado en nuevo Workbook
    wb_src = load_workbook(io.BytesIO(file_bytes))
    ws_src = wb_src[sheet_name] if sheet_name and sheet_name in wb_src.sheetnames else wb_src.active

    headers = [str(cell.value) if cell.value is not None else '' for cell in ws_src[1]]
    rows_removed = 0

    # Primera pasada — eliminar hipervínculos en todo el archivo
    hyperlinks_removed = 0
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.hyperlink is not None:
                cell.hyperlink = None
                hyperlinks_removed += 1

    # Leer todas las filas aplicando correcciones de correo y reglas de formato
    data_rows = []
    for data_idx, row in enumerate(ws_src.iter_rows(min_row=2, values_only=False)):
        # Saltar filas marcadas para eliminar
        if data_idx in email_rows_to_delete:
            rows_removed += 1
            continue

        new_row = []
        for col_0idx, cell in enumerate(row):
            col_name = headers[col_0idx] if col_0idx < len(headers) else ''
            rule     = rules.get(col_name, 'none')
            val      = cell.value

            # Aplicar corrección de correo (antes del casing)
            email_fix = email_fixes_by_col.get(col_name, {}).get(data_idx)
            if email_fix is not None:
                val = email_fix
                emails_fixed += 1

            if rule != 'none':
                if val is not None:
                    val_str = str(val) if not isinstance(val, str) else val
                    val = clean_rut(val_str) if rule == 'clean_rut' else apply_casing(val_str, rule)
                    val = remove_accents(val)
            new_row.append(val)
        data_rows.append(new_row)

    # Escribir en nuevo Workbook
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.append(headers)
    for row in data_rows:
        ws_new.append(row)

    valid_indices   = [i for i, h in enumerate(headers) if h]
    preview_columns = [headers[i] for i in valid_indices]
    preview_rows    = [
        [str(r[i]) if i < len(r) and r[i] is not None else '' for i in valid_indices]
        for r in data_rows[:5]
    ]

    out = io.BytesIO()
    wb_new.save(out)
    out.seek(0)
    log_event(
        action='confirm',
        ip=request.remote_addr,
        files=[{'name': filename, 'size_kb': 0}],
        mode='formatear',
        rows=len(data_rows),
    )
    return jsonify({
        'output_data':        base64.b64encode(out.read()).decode('utf-8'),
        'hyperlinks_removed': hyperlinks_removed,
        'emails_fixed':       emails_fixed,
        'rows_removed':       rows_removed,
        'preview_columns':    preview_columns,
        'preview_rows':       preview_rows,
    })


if __name__ == '__main__':
    import os
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 8080)))
